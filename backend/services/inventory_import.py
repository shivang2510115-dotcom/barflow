"""Reading a stock file, and saying exactly what it would do — without doing any of it.

A property opening with two hundred lines of stock does not type them in, so the stock
figures are wrong from the first day and every low-stock alert built on them is noise.
The fix is to accept the file the supplier already sent. The risk is that the file is a
supplier's file: misspelt names, the same item twice, quantities in cases where this
system counts bottles, and prices with a rupee sign glued to them. Writing that straight
into stock produces numbers nobody trusts, and stock nobody trusts is worse than stock
nobody entered.

So this module **only reads**. It turns bytes into a report — for each row, what the
import would do and what it could not understand — and a second function turns rows a
human has since corrected into the writes they mean. Neither one touches the database;
`routers/inventory.py` executes the plan and is the only place that writes. That split is
what makes the whole feature testable without a server, and it is why the review step in
the middle can exist at all.

**Two rules worth stating out loud.**

*Matching* is on the item name, case- and whitespace-insensitive: the name is stripped,
its internal runs of whitespace collapse to one space, and it is casefolded. So
`"  BOURBON   750ML "` is the item already stored as `"Bourbon 750ml"`. Nothing cleverer
— no fuzzy distance, no stemming — because a near-match that silently picks the wrong
item is precisely the outcome this feature exists to prevent, and the review screen is
where a human resolves the ones a rule cannot. Note that `"Gin 750 ml"` and `"Gin 750ml"`
are therefore *different* items; the admin re-points one at the other by hand if they
meant the same thing.

*Quantity replaces, it does not add.* An update showing `8 → 20` sets the stock to 20.
This is a stock take, not a delivery note: the file says what is on the shelf. Adding
would make a file uploaded twice double the whole store room, which is the kind of
mistake nobody notices until the numbers are already wrong.
"""
import csv
import io
import re

# What one upload may weigh. This deploys as a Firebase Function, where an unbounded
# upload is a way to be billed for someone else's afternoon — and the guard has to be a
# number, not a hope, because the caller controls the file. 1 MB is roughly twenty
# thousand rows of stock; the two-hundred-line opening this feature was built for is
# about ten kilobytes, so the bound is two orders of magnitude clear of the real case
# and still small enough that a function instance can hold it in memory without thinking.
MAX_UPLOAD_BYTES = 1024 * 1024
MAX_UPLOAD_LABEL = "1 MB"

# And a second bound, on rows rather than bytes, because the two fail differently: a
# small file of very many short rows passes the byte check and would still produce a
# review screen no human can review and a response no browser enjoys. A store room with
# more than two thousand distinct lines is doing something this screen is the wrong tool
# for.
MAX_ROWS = 2000

# The columns the template publishes, in the order it publishes them.
TEMPLATE_COLUMNS = ("name", "unit", "stock", "threshold", "cost_per_unit", "category")

# Of those, the ones the file must actually carry. `threshold`, `cost_per_unit` and
# `category` are optional on purpose: a supplier's sheet has a name, a unit, a quantity
# and a price, and it has never heard of a low-stock threshold. Refusing the file over a
# column that only exists inside BarFlow would be refusing the ordinary case, so a
# missing optional column falls back to the same default a hand-typed item gets.
REQUIRED_COLUMNS = ("name", "unit", "stock")

# Defaults, kept identical to routers.inventory.InventoryItemIn. Two lists that are meant
# to agree and are written down twice is how they stop agreeing, so the router builds the
# real item through its own model and these only fill blank cells.
DEFAULT_THRESHOLD = 5.0
DEFAULT_COST = 0.0
DEFAULT_CATEGORY = "spirits"

# What a header cell may say and still be understood. Spreadsheets arrive from suppliers,
# accountants and whoever exported them, so "Qty", "UOM" and "Reorder Level" all turn up.
# Keys here are already normalised (see `_normalise_header`), so case, padding, hyphens
# and underscores are handled before the lookup rather than multiplied inside it.
COLUMN_ALIASES: dict[str, str] = {
    "name": "name", "item": "name", "item name": "name", "product": "name",
    "product name": "name", "description": "name", "particulars": "name",

    "unit": "unit", "units": "unit", "uom": "unit", "unit of measure": "unit",
    "measure": "unit", "pack": "unit",

    "stock": "stock", "qty": "stock", "quantity": "stock", "current stock": "stock",
    "current quantity": "stock", "opening stock": "stock", "on hand": "stock",
    "in stock": "stock", "closing stock": "stock", "count": "stock",

    "threshold": "threshold", "low stock threshold": "threshold",
    "low stock": "threshold", "reorder level": "threshold", "reorder point": "threshold",
    "alert at": "threshold", "min": "threshold", "minimum": "threshold",
    "min stock": "threshold", "par": "threshold", "par level": "threshold",

    "cost per unit": "cost_per_unit", "cost": "cost_per_unit",
    "unit cost": "cost_per_unit", "price": "cost_per_unit",
    "unit price": "cost_per_unit", "rate": "cost_per_unit",
    "purchase price": "cost_per_unit", "cost price": "cost_per_unit",

    "category": "category", "type": "category", "group": "category",
    "class": "category", "section": "category",
}

# The order cells are checked in, so a row with two bad cells reports them in the order
# somebody reading the file would meet them rather than in dictionary order.
_CELL_ORDER = ("name", "unit", "stock", "threshold", "cost_per_unit", "category")

# The money and quantity noise a real sheet carries. Stripped rather than refused: a
# price written "₹ 3,500.50" or "Rs 60/-" is a price, and making the owner clean their
# own file before the tool will read it is making the tool useless.
_CURRENCY_PREFIX = re.compile(r"^(?:rs\.?|inr|₹|₹)\s*", re.IGNORECASE)
_TRAILING_DASH = re.compile(r"/-\s*$")


class ImportRefused(Exception):
    """The upload is not something this reader can open at all.

    Distinct from the per-row and per-file problems reported *inside* a plan: those are
    the report doing its job, and the owner sees them on the review screen. This is the
    request being refused before there is anything to review — the file is too big, or it
    is a PDF. The router turns it into the HTTP status named here.
    """

    def __init__(self, status_code: int, detail: str):
        super().__init__(detail)
        self.status_code = status_code
        self.detail = detail


# ------------------------------- the template -------------------------------
def template_csv() -> str:
    """The file to fill in: the exact columns, and one example row.

    Published rather than inferred. Guessing a format is where a bulk import dies — the
    owner exports something plausible, the reader refuses it, and nobody tries twice — so
    the shape is a download, and `test_the_template_is_a_file_this_parser_accepts` holds
    the template and the reader to each other.
    """
    return (",".join(TEMPLATE_COLUMNS) + "\n"
            + "Bourbon 750ml,bottle,12,4,3500,spirits\n")


# ------------------------------- the reading -------------------------------
def match_key(name: str) -> str:
    """The name, reduced to what two spellings of one item have in common.

    See the module docstring: strip, collapse internal whitespace, casefold. Deliberately
    nothing else.
    """
    return " ".join(str(name or "").split()).casefold()


def _clean(value) -> str:
    """One cell as text, with the padding gone and internal runs of space collapsed."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _normalise_header(value) -> str:
    return " ".join(str(value or "").replace("_", " ").replace("-", " ").split()).lower()


def _to_number(text: str):
    """A number from a cell a person typed, or None if it is not one.

    None means "unreadable", and the caller turns that into an error naming the cell.
    An empty cell is *not* unreadable — it is absent — so the caller checks for that
    first and either defaults it or refuses it, depending on the column.
    """
    cleaned = _CURRENCY_PREFIX.sub("", text.strip())
    cleaned = _TRAILING_DASH.sub("", cleaned)
    cleaned = cleaned.replace("₹", "").replace("₹", "")
    cleaned = cleaned.replace(",", "").replace(" ", "").replace(" ", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _decode(raw: bytes) -> str:
    """Bytes to text, trying the encodings a spreadsheet actually exports as.

    `utf-8-sig` first, which is plain UTF-8 plus the byte-order mark Excel writes — left
    on, the mark glues itself to the first header cell and `name` stops being recognised
    as a column. Then cp1252, which is what a Windows Excel "CSV" usually is and which is
    where a stray `₹` or a smart quote comes from. latin-1 last because it cannot fail,
    so an odd file is read imperfectly and reported on rather than refused outright.
    """
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def _grid_from_csv(text: str) -> list[tuple[int, list[str]]]:
    """(line number, cells) for every line, blank ones included.

    The line number comes from the reader rather than from an enumerate, so a quoted
    field containing a newline does not silently shift every row number below it — and
    the number an error names is the one the owner can scroll to in their spreadsheet.
    """
    reader = csv.reader(io.StringIO(text, newline=""))
    grid = []
    for cells in reader:
        grid.append((reader.line_num, [str(c) for c in cells]))
    return grid


def _grid_from_xlsx(raw: bytes) -> list[tuple[int, list[str]]]:
    """The same, from the first sheet of a workbook.

    openpyxl is imported here rather than at module scope for two reasons: this deploys
    as a Firebase Function, where every import is cold-start latency paid on requests that
    are not this one; and if the dependency is ever dropped from the image, a missing
    Excel reader must degrade to "export a CSV instead" rather than take the whole
    inventory router down at import.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportRefused(400, (
            "This server cannot read Excel files. Open the sheet in Excel or Google "
            "Sheets and export it as CSV (.csv), then upload that.")) from None

    try:
        book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise ImportRefused(400, (
            "That .xlsx file could not be opened. If it was renamed from another format, "
            "export it again as CSV (.csv) and upload that.")) from None

    grid = []
    try:
        sheet = book.worksheets[0]
        for number, cells in enumerate(sheet.iter_rows(values_only=True), start=1):
            # A spreadsheet holds numbers as numbers, so 3500 arrives as an int and never
            # as the string the CSV path would see. Rendering an integral float as "12"
            # rather than "12.0" only matters for how a value is echoed back in an error
            # message, but that message is the whole product of a bad cell.
            row = []
            for cell in cells:
                if cell is None:
                    row.append("")
                elif isinstance(cell, float) and cell.is_integer():
                    row.append(str(int(cell)))
                else:
                    row.append(str(cell))
            grid.append((number, row))
            if number > MAX_ROWS + 2:
                break
    finally:
        book.close()
    return grid


def plan_upload(raw: bytes, filename: str, existing: list[dict]) -> dict:
    """Read an uploaded file and report what importing it would do. Writes nothing.

    `existing` is this property's current stock, read by the caller through its scoped
    handle — passed in rather than fetched here so that this stays a pure function over
    its arguments and so that matching can never reach another hotel's items.
    """
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ImportRefused(413, (
            f"That file is larger than {MAX_UPLOAD_LABEL}, which is the most this "
            f"import will read. A stock list of a few thousand lines is well under it — "
            f"if the file is big because it carries images or extra sheets, export just "
            f"the stock rows as CSV."))

    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension in ("csv", "txt", ""):
        grid = _grid_from_csv(_decode(raw))
    elif extension in ("xlsx", "xlsm"):
        grid = _grid_from_xlsx(raw)
    elif extension == "xls":
        raise ImportRefused(400, (
            "That is the old .xls format, which this import cannot read. Open it and "
            "save as .xlsx, or export it as CSV (.csv)."))
    else:
        raise ImportRefused(400, (
            f"This import reads CSV (.csv) and Excel (.xlsx) files, and that one is "
            f"'.{extension}'. Export your stock list as .csv and upload it again."))

    return build_report(grid, existing)


def build_report(grid: list[tuple[int, list[str]]], existing: list[dict]) -> dict:
    """The plan itself: a header, a row for each line of stock, and a summary."""
    file_errors: list[str] = []

    header_index = next((i for i, (_n, cells) in enumerate(grid)
                         if any(_clean(c) for c in cells)), None)
    if header_index is None:
        return _report([], ["The file is empty — there is nothing in it to import. "
                            "Download the template to see the columns and an example "
                            "row."], {})

    _header_line, header_cells = grid[header_index]
    columns: dict[str, int] = {}
    for position, cell in enumerate(header_cells):
        canonical = COLUMN_ALIASES.get(_normalise_header(cell))
        if canonical and canonical not in columns:
            columns[canonical] = position

    missing = [c for c in REQUIRED_COLUMNS if c not in columns]
    if missing:
        named = ", ".join(missing)
        return _report([], [
            f"This file is missing the {named} column"
            f"{'s' if len(missing) > 1 else ''}, which the import cannot do without. "
            f"Download the template to see the columns and an example row."], columns)

    body = [(number, cells) for number, cells in grid[header_index + 1:]
            if any(_clean(c) for c in cells)]
    if not body:
        return _report([], ["The file has a header but no rows of stock below it."],
                       columns)
    if len(body) > MAX_ROWS:
        return _report([], [
            f"That file has {len(body)} rows and this import reads at most {MAX_ROWS} "
            f"at a time. Split it and upload it in parts."], columns)

    existing_by_key = {match_key(item["name"]): item for item in existing}
    first_seen: dict[str, int] = {}
    rows = []

    for number, cells in body:
        def cell(name: str) -> str:
            position = columns.get(name)
            if position is None or position >= len(cells):
                return ""
            return cells[position]

        values, errors = _read_cells({name: cell(name) for name in _CELL_ORDER}, number)

        key = match_key(values["name"])
        kind, item_id, before, duplicate_of = "new", None, None, None

        if key and key in first_seen:
            kind = "duplicate"
            duplicate_of = first_seen[key]
            errors.append(_error(number, "name", values["name"], (
                f"the same item is already on row {duplicate_of} — drop one of the two "
                f"rows, or merge them into a single quantity")))
        elif key and key in existing_by_key:
            match = existing_by_key[key]
            kind = "update"
            item_id = match["id"]
            before = _snapshot(match)
            # The stored spelling wins over the supplier's. An import is not a rename:
            # the owner named this item, and a sheet shouting "GIN 750ML" must not
            # silently retitle it on the shelf. Renaming stays on the Inventory screen.
            values["name"] = match["name"]

        if key and key not in first_seen:
            first_seen[key] = number

        rows.append({
            "row": number,
            **values,
            "kind": kind,
            # A row nobody can read is not applied until somebody has looked at it. The
            # review screen may set this to create or update once the cell is fixed.
            "action": "skip" if errors else ("update" if kind == "update" else "create"),
            "item_id": item_id,
            "existing": before,
            "duplicate_of": duplicate_of,
            "errors": errors,
        })

    return _report(rows, file_errors, columns)


def _read_cells(cells: dict[str, str], number: int) -> tuple[dict, list[dict]]:
    """One row's cells, parsed, plus an error for each one that could not be."""
    values: dict = {}
    errors: list[dict] = []

    name = _clean(cells["name"])
    if not name:
        errors.append(_error(number, "name", cells["name"],
                             "this row has no item name, so there is nothing to import"))
    values["name"] = name

    # Lower-cased so that "Bottle" and "bottle" are one unit rather than two lines in the
    # store room. Not validated against a list: a supplier counts in cases, crates and
    # pieces, and the point of the review screen is that a human sees "case" next to a
    # system that counts bottles and fixes it — a refusal here would only send them back
    # to edit the spreadsheet.
    unit = _clean(cells["unit"]).lower()
    if not unit:
        errors.append(_error(number, "unit", cells["unit"], (
            "this row has no unit, and a case is not a bottle — say which one this "
            "quantity is counted in")))
    values["unit"] = unit

    values["stock"] = _number_cell(cells["stock"], "stock", number, errors,
                                   default=None, quantity=True)
    values["threshold"] = _number_cell(cells["threshold"], "threshold", number, errors,
                                       default=DEFAULT_THRESHOLD, quantity=True)
    values["cost_per_unit"] = _number_cell(cells["cost_per_unit"], "cost_per_unit",
                                           number, errors, default=DEFAULT_COST,
                                           quantity=False)
    values["category"] = _clean(cells["category"]).lower() or DEFAULT_CATEGORY
    return values, errors


def _number_cell(raw: str, column: str, number: int, errors: list, *, default,
                 quantity: bool):
    """One numeric cell. `default=None` marks the column as one that must be filled in."""
    text = str(raw or "").strip()
    if not text:
        if default is None:
            errors.append(_error(number, column, raw, (
                "this row has no quantity — a blank is not zero, and a stock figure "
                "nobody entered is the problem this import exists to fix")))
            return 0.0
        return float(default)

    value = _to_number(text)
    if value is None:
        errors.append(_error(number, column, text,
                             f"'{text}' is not a number this import can read"))
        return 0.0
    if value < 0:
        noun = "a quantity" if quantity else "a cost"
        errors.append(_error(number, column, text, f"'{text}' is negative, and {noun} "
                                                   f"cannot be"))
        return 0.0
    return value


def _error(number: int, column: str, value, message: str) -> dict:
    """A problem, named by the row and the cell it is in.

    "Invalid file" tells the owner nothing; "row 14, column stock" tells them where to
    look. The row number, the column and the original text are all carried separately as
    well as inside the sentence, so the review screen can highlight the cell rather than
    only print the sentence.
    """
    return {"row": number, "column": column, "value": str(value or "").strip(),
            "message": f"row {number}, column {column}: {message}"}


def _snapshot(item: dict) -> dict:
    return {"id": item.get("id"), "name": item.get("name"), "unit": item.get("unit"),
            "stock": item.get("stock"), "threshold": item.get("threshold"),
            "cost_per_unit": item.get("cost_per_unit"),
            "category": item.get("category")}


def _report(rows: list, file_errors: list, columns: dict) -> dict:
    return {
        "rows": rows,
        "file_errors": file_errors,
        "columns": sorted(columns),
        "summary": {
            "total": len(rows),
            "new": sum(1 for r in rows if r["kind"] == "new"),
            "update": sum(1 for r in rows if r["kind"] == "update"),
            "duplicate": sum(1 for r in rows if r["kind"] == "duplicate"),
            "blocked": sum(1 for r in rows if r["errors"]),
        },
        "limits": {"max_bytes": MAX_UPLOAD_BYTES, "max_label": MAX_UPLOAD_LABEL,
                   "max_rows": MAX_ROWS},
    }


# ------------------------------- the writing -------------------------------
def plan_apply(rows: list[dict], existing: list[dict]) -> tuple[list[dict], list[dict]]:
    """Turn rows a human has reviewed into the writes they mean.

    Returns `(operations, refusals)`. **Never both**: if anything at all is refused, no
    operation is returned, because a partial write nobody asked for is worse than a
    refusal they can act on. The router turns a non-empty refusal list into a 400 and
    writes nothing.

    Everything the review screen claims is re-decided here rather than trusted. The rows
    arrive from a browser, which means they may have been edited since the preview — that
    is the whole point of the screen — but also that the classification, the item id and
    the numbers in them are all client input. So the values are parsed again by the same
    rules, an update has to name an item this property still holds, and a create whose
    name already exists is refused rather than quietly duplicating an item.
    """
    operations: list[dict] = []
    refusals: list[dict] = []

    by_id = {item["id"]: item for item in existing}
    by_key = {match_key(item["name"]): item for item in existing}
    claimed_keys: dict[str, int] = {}
    claimed_ids: dict[str, int] = {}

    def refuse(number, name, message):
        refusals.append({"row": number, "name": name, "message": message})

    for row in rows:
        number = row.get("row") or 0
        action = str(row.get("action") or "skip").strip().lower()
        if action == "skip":
            continue
        if action not in ("create", "update"):
            refuse(number, _clean(row.get("name")),
                   f"row {number}: '{action}' is not something this import can do")
            continue

        values, errors = _read_cells(
            {name: str(row.get(name, "") if row.get(name) is not None else "")
             for name in _CELL_ORDER}, number)
        if errors:
            refuse(number, values["name"], "; ".join(e["message"] for e in errors))
            continue

        if action == "update":
            item = by_id.get(row.get("item_id") or "")
            if not item:
                refuse(number, values["name"], (
                    f"row {number}: the stock item this row updates is no longer in your "
                    f"inventory — it may have been deleted while this file was open. "
                    f"Upload the file again."))
                continue
            if item["id"] in claimed_ids:
                refuse(number, values["name"], (
                    f"row {number}: row {claimed_ids[item['id']]} already updates "
                    f"'{item['name']}' — two rows cannot both set one item's stock"))
                continue
            claimed_ids[item["id"]] = number
            operations.append({
                "row": number, "action": "update", "item_id": item["id"],
                # The stored name stands; see build_report for why an import is not a
                # rename.
                "item": {**values, "name": item["name"]},
                "before": _snapshot(item),
            })
            continue

        key = match_key(values["name"])
        if key in by_key:
            refuse(number, values["name"], (
                f"row {number}: '{by_key[key]['name']}' is already in your inventory. "
                f"Set this row to update it instead of creating a second copy."))
            continue
        if key in claimed_keys:
            refuse(number, values["name"], (
                f"row {number}: '{values['name']}' is also created on row "
                f"{claimed_keys[key]} — drop one of the two rows"))
            continue
        claimed_keys[key] = number
        operations.append({"row": number, "action": "create", "item_id": None,
                           "item": values, "before": None})

    if refusals:
        return [], refusals
    return operations, refusals
